"""Optimiertes Excel-Layout für den Dienstplan der SBP.

4 Sheets:
1. Dienstplan – Wochen-Kalenderansicht mit Farbcodierung + TVK-Ampel
2. Tagesansicht – Detaillierte Liste aller Dienste
3. TVK-Prüfbericht – Violations mit Vorschlägen
4. Statistik – Dashboard-Übersicht
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter

from ..models.plan import Dienstplan
from ..models.events import DienstType, Dienst
from ..constraints.base import Violation, Severity


# Farben
COLORS = {
    "header_bg": "2C3E50",       # Dunkelblau
    "header_font": "FFFFFF",      # Weiß
    "frei": "2ECC71",            # Grün
    "urlaub": "27AE60",          # Dunkelgrün
    "probe": "3498DB",           # Blau
    "konzert": "E74C3C",         # Rot
    "tour": "E67E22",            # Orange
    "sk": "F39C12",              # Gelb
    "sonstiges": "95A5A6",       # Grau
    "rza": "BDC3C7",             # Hellgrau
    "tvk_ok": "2ECC71",          # Grün
    "tvk_warn": "F39C12",        # Gelb
    "tvk_error": "E74C3C",       # Rot
    "error_bg": "FADBD8",        # Helles Rot
    "warning_bg": "FEF9E7",      # Helles Gelb
    "info_bg": "EBF5FB",         # Helles Blau
    "even_row": "F8F9FA",        # Wechselzeile
    "border": "DEE2E6",          # Rahmen
}

THIN_BORDER = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=COLORS["header_font"])
HEADER_FILL = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

DIENST_TYPE_COLORS = {
    DienstType.FREI: COLORS["frei"],
    DienstType.URLAUB: COLORS["urlaub"],
    DienstType.REISEZEITAUSGLEICH: COLORS["rza"],
    DienstType.PROBE: COLORS["probe"],
    DienstType.GENERALPROBE: COLORS["probe"],
    DienstType.HAUPTPROBE: COLORS["probe"],
    DienstType.ANSPIELPROBE: COLORS["probe"],
    DienstType.KONZERT: COLORS["konzert"],
    DienstType.ABO_KONZERT: COLORS["konzert"],
    DienstType.GASTSPIEL: COLORS["konzert"],
    DienstType.SCHUELERKONZERT: COLORS["sk"],
    DienstType.BABYKONZERT: COLORS["sk"],
    DienstType.DIRIGIERKURS: COLORS["sk"],
    DienstType.REISE: COLORS["tour"],
    DienstType.PODCAST: COLORS["sonstiges"],
    DienstType.TONAUFNAHME: COLORS["sonstiges"],
    DienstType.SONSTIGES: COLORS["sonstiges"],
}


def write_dienstplan(plan: Dienstplan, output_path: str | Path):
    """Schreibt den Dienstplan als optimiertes Excel."""
    wb = Workbook()

    _write_calendar_sheet(wb, plan)
    _write_detail_sheet(wb, plan)
    _write_violations_sheet(wb, plan)
    _write_statistics_sheet(wb, plan)

    # Default-Sheet entfernen falls vorhanden
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(str(output_path))


def _write_calendar_sheet(wb: Workbook, plan: Dienstplan):
    """Sheet 1: Wochen-Kalenderansicht."""
    ws = wb.active
    ws.title = "Dienstplan"

    # Titel
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Dienstplan {plan.orchestra_name} | {plan.plan_start.strftime('%d.%m.%Y')} – {plan.plan_end.strftime('%d.%m.%Y')}"
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header
    headers = ["KW", "Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag", "Dienste", "TVK"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 6    # KW
    for col in range(2, 9):                # Mo-So
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.column_dimensions["I"].width = 8    # Dienste
    ws.column_dimensions["J"].width = 6    # TVK

    # Wochen-Daten
    row = 4
    for week in plan.weeks:
        ws.row_dimensions[row].height = 50

        # KW
        cell = ws.cell(row=row, column=1, value=f"KW {week.week_number}")
        cell.font = Font(name="Calibri", size=9, bold=True)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

        # Tage Mo-So
        for day_offset in range(7):
            day_date = week.start_date + timedelta(days=day_offset)
            col = day_offset + 2
            dienst = week.dienst_for_date(day_date)

            if dienst:
                text = f"{day_date.strftime('%d.%m.')}\n{dienst.summary}"
                if dienst.dienst_count > 0:
                    text += f"\n[{dienst.dienst_count:g}]"

                cell = ws.cell(row=row, column=col, value=text)

                # Farbcodierung
                color = _get_dienst_color(dienst)
                if color:
                    cell.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )
                    # Helle Schrift auf dunklem Hintergrund
                    cell.font = Font(name="Calibri", size=9, color="FFFFFF")
                else:
                    cell.font = Font(name="Calibri", size=9)
            else:
                cell = ws.cell(row=row, column=col, value=day_date.strftime('%d.%m.'))
                cell.font = Font(name="Calibri", size=9, color="999999")

            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        # Dienste-Summe
        total = week.total_dienste
        cell = ws.cell(row=row, column=9, value=total)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        cell.number_format = "0.#"

        # TVK-Ampel
        tvk_cell = ws.cell(row=row, column=10)
        if week.tvk_status == "error":
            tvk_cell.value = "!!!"
            tvk_cell.fill = PatternFill(start_color=COLORS["tvk_error"], end_color=COLORS["tvk_error"], fill_type="solid")
            tvk_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        elif week.tvk_status == "warning":
            tvk_cell.value = "!"
            tvk_cell.fill = PatternFill(start_color=COLORS["tvk_warn"], end_color=COLORS["tvk_warn"], fill_type="solid")
            tvk_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        else:
            tvk_cell.value = "OK"
            tvk_cell.fill = PatternFill(start_color=COLORS["tvk_ok"], end_color=COLORS["tvk_ok"], fill_type="solid")
            tvk_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        tvk_cell.alignment = CENTER_ALIGN
        tvk_cell.border = THIN_BORDER

        row += 1

    # Zusammenfassung
    row += 1
    ws.cell(row=row, column=1, value="Gesamt:").font = Font(bold=True)
    ws.cell(row=row, column=9, value=plan.total_dienste).font = Font(bold=True, size=12)
    ws.cell(row=row, column=9).number_format = "0.#"


def _write_detail_sheet(wb: Workbook, plan: Dienstplan):
    """Sheet 2: Detaillierte Tagesansicht."""
    ws = wb.create_sheet("Tagesansicht")

    headers = ["Datum", "Tag", "Zeit", "Typ", "Formation", "Programm", "Ort", "Leitung", "Dienste", "Bemerkungen"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Spaltenbreiten
    widths = [12, 4, 16, 14, 18, 30, 14, 12, 8, 30]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    row = 2
    for dienst in plan.all_dienste():
        if not dienst.events and dienst.dienst_count == 0:
            # Freie Tage auch anzeigen
            _write_detail_row(ws, row, dienst, None, row % 2 == 0)
            row += 1
        else:
            for event in dienst.events:
                _write_detail_row(ws, row, dienst, event, row % 2 == 0)
                row += 1
            if not dienst.events:
                _write_detail_row(ws, row, dienst, None, row % 2 == 0)
                row += 1

    # Autofilter
    ws.auto_filter.ref = f"A1:J{row - 1}"


def _write_detail_row(ws, row: int, dienst: Dienst, event, is_even: bool):
    """Schreibt eine Zeile in die Tagesansicht."""
    from ..models.events import Event

    if is_even:
        fill = PatternFill(start_color=COLORS["even_row"], end_color=COLORS["even_row"], fill_type="solid")
    else:
        fill = PatternFill(fill_type=None)

    ws.cell(row=row, column=1, value=dienst.dienst_date.strftime("%d.%m.%Y"))
    ws.cell(row=row, column=2, value=dienst.day_of_week)

    if event:
        zeit = ""
        if event.start_time:
            zeit = event.start_time.strftime("%H:%M")
            if event.end_time:
                zeit += f" - {event.end_time.strftime('%H:%M')}"
        ws.cell(row=row, column=3, value=zeit)
        ws.cell(row=row, column=4, value=event.dienst_type.value)
        ws.cell(row=row, column=5, value=event.formation.value if event.formation else "")
        ws.cell(row=row, column=6, value=event.programm)
        ws.cell(row=row, column=7, value=event.ort)
        ws.cell(row=row, column=8, value=event.leitung)
        ws.cell(row=row, column=10, value=event.sonstiges)
    else:
        ws.cell(row=row, column=4, value="Frei" if dienst.is_free else "")

    ws.cell(row=row, column=9, value=dienst.dienst_count)
    ws.cell(row=row, column=9).number_format = "0.#"

    for col in range(1, 11):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGN
        cell.border = THIN_BORDER
        if fill.fill_type:
            cell.fill = fill


def _write_violations_sheet(wb: Workbook, plan: Dienstplan):
    """Sheet 3: TVK-Prüfbericht."""
    ws = wb.create_sheet("TVK-Prüfbericht")

    headers = ["Schwere", "Regel", "§ TVK", "KW / Datum", "Ist", "Grenzwert", "Beschreibung", "Vorschlag"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    widths = [10, 22, 10, 14, 8, 10, 40, 35]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    row = 2
    for v in plan.violations:
        bg_color = {
            Severity.ERROR: COLORS["error_bg"],
            Severity.WARNING: COLORS["warning_bg"],
            Severity.INFO: COLORS["info_bg"],
        }.get(v.severity, "FFFFFF")
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        ws.cell(row=row, column=1, value=v.severity_icon)
        ws.cell(row=row, column=2, value=v.rule_name)
        ws.cell(row=row, column=3, value=v.tvk_paragraph)
        if v.affected_week:
            ws.cell(row=row, column=4, value=f"KW {v.affected_week}")
        elif v.affected_dates:
            ws.cell(row=row, column=4, value=v.affected_dates[0].strftime("%d.%m."))
        ws.cell(row=row, column=5, value=v.current_value)
        ws.cell(row=row, column=6, value=v.limit_value)
        ws.cell(row=row, column=7, value=v.message)
        ws.cell(row=row, column=8, value=v.suggestion)

        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER
            cell.fill = row_fill

        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="Keine Verstöße gefunden").font = Font(italic=True, color="27AE60")

    ws.auto_filter.ref = f"A1:H{max(row - 1, 2)}"


def _write_statistics_sheet(wb: Workbook, plan: Dienstplan):
    """Sheet 4: Statistik-Dashboard."""
    ws = wb.create_sheet("Statistik")

    # Titel
    ws.merge_cells("A1:D1")
    ws["A1"].value = "Statistik-Übersicht"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 15

    # Kennzahlen
    stats = [
        ("Zeitraum", f"{plan.plan_start.strftime('%d.%m.%Y')} – {plan.plan_end.strftime('%d.%m.%Y')}"),
        ("Kalenderwochen", plan.total_weeks),
        ("Gesamt-Dienste", plan.total_dienste),
        ("Durchschnitt Dienste/Woche", plan.avg_dienste_per_week),
        ("Freie Tage", plan.total_free_days),
        ("Freie Sonntage", plan.free_sundays),
        ("Wochen mit TVK-Verstoß", plan.weeks_with_violations),
    ]

    row = 3
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = BODY_FONT
        if isinstance(value, float):
            cell.number_format = "0.#"
        row += 1

    # Violations-Zusammenfassung
    row += 1
    ws.cell(row=row, column=1, value="TVK-Prüfergebnis").font = Font(size=12, bold=True)
    row += 1

    errors = sum(1 for v in plan.violations if v.severity == Severity.ERROR)
    warnings = sum(1 for v in plan.violations if v.severity == Severity.WARNING)
    infos = sum(1 for v in plan.violations if v.severity == Severity.INFO)

    for label, count, color in [
        ("Fehler (Verstöße)", errors, COLORS["tvk_error"]),
        ("Warnungen", warnings, COLORS["tvk_warn"]),
        ("Hinweise", infos, COLORS["tvk_ok"]),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=count)
        cell.font = Font(bold=True, color=color)
        row += 1

    # Dienste-Verteilung nach Typ
    row += 1
    ws.cell(row=row, column=1, value="Dienste nach Typ").font = Font(size=12, bold=True)
    row += 1

    by_type = plan.dienste_by_type()
    for dtype, count in sorted(by_type.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=dtype.value).font = BODY_FONT
        ws.cell(row=row, column=2, value=count).font = BODY_FONT
        row += 1

    # Wochen-Übersicht
    row += 1
    ws.cell(row=row, column=4, value="Dienste pro Woche").font = Font(size=12, bold=True)
    row += 1
    ws.cell(row=row, column=4, value="KW").font = Font(bold=True)
    ws.cell(row=row, column=5, value="Dienste").font = Font(bold=True)
    row += 1

    for week in plan.weeks:
        ws.cell(row=row, column=4, value=f"KW {week.week_number}")
        cell = ws.cell(row=row, column=5, value=week.total_dienste)
        cell.number_format = "0.#"
        if week.tvk_status == "error":
            cell.font = Font(bold=True, color=COLORS["tvk_error"])
        elif week.tvk_status == "warning":
            cell.font = Font(color=COLORS["tvk_warn"])
        row += 1


def _get_dienst_color(dienst: Dienst) -> str | None:
    """Bestimmt die Hintergrundfarbe für einen Dienst in der Kalenderansicht."""
    if dienst.is_free:
        primary = dienst.primary_type
        return DIENST_TYPE_COLORS.get(primary, COLORS["frei"])
    if dienst.dienst_count == 0:
        return COLORS["rza"]
    return DIENST_TYPE_COLORS.get(dienst.primary_type)
