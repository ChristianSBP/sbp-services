"""Liest den Jahresplan der SBP aus einer Excel-Datei.

Das Excel-Format der SBP:
- Sheet "Tabelle1"
- Zeile 1: Monatsüberschriften (in der Inhalt-Spalte jedes Monats)
- Spalte A: Wochentag-Kürzel (Mo, Di, Mi, Do, Fr, Sa, So) — nur visuell
- Je 2 Spalten pro Monat: (Tagesnummer-Spalte, Inhalt-Spalte)
- Jeder Monat hat seine eigene Tagesnummern-Spalte (Zeilen variieren!)
- Spalten-Mapping (automatisch erkannt):
    Jan=B/C, Feb=D/E, Mär=F/G, Apr=H/I, Mai=J/K, Jun=L/M
    Jul=O/P, Aug=Q/R, Sep=S/T, Okt=U/V, Nov=W/X, Dez=Y/Z
  Hinweis: Spalte N wird übersprungen (zwischen Juni und Juli).
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.cell import Cell


MONTH_NAMES = [
    "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember",
]

DAY_ABBREVS = {"Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"}


class JahresplanCell:
    """Eine extrahierte Zelle aus dem Jahresplan."""

    def __init__(self, month: int, day: int, year: int, text: str, cell_color: Optional[str] = None):
        self.month = month
        self.day = day
        self.year = year
        self.text = text.strip() if text else ""
        self.cell_color = cell_color

    @property
    def event_date(self) -> date:
        return date(self.year, self.month, self.day)

    def __repr__(self) -> str:
        return f"JahresplanCell({self.event_date}, {self.text!r})"


def read_jahresplan(filepath: str | Path, year: int = 2026) -> List[JahresplanCell]:
    """Liest alle Events aus dem Jahresplan-Excel.

    Returns:
        Liste von JahresplanCell mit Datum und Rohtext pro Event.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb[wb.sheetnames[0]]  # Erstes Sheet (Tabelle1)

    # Schritt 1: Monats-Spalten finden (Tages-Spalte + Inhalt-Spalte pro Monat)
    month_columns = _find_month_columns(ws)

    # Schritt 2+3: Pro Monat eigene Tagesnummern scannen + Events extrahieren
    cells: List[JahresplanCell] = []
    for month_num, (day_col, event_col) in month_columns.items():
        # Jeder Monat hat seine eigene Tagesnummern-Spalte!
        day_rows = _find_day_rows_for_column(ws, day_col)

        for row_num, day_num in day_rows.items():
            # Prüfe ob dieser Tag im Monat existiert
            try:
                d = date(year, month_num, day_num)
            except ValueError:
                continue

            cell = ws.cell(row=row_num, column=event_col)
            text = _get_cell_text(cell)
            if text:
                color = _get_cell_color(cell)
                cells.append(JahresplanCell(
                    month=month_num,
                    day=day_num,
                    year=year,
                    text=text,
                    cell_color=color,
                ))

    wb.close()
    return sorted(cells, key=lambda c: c.event_date)


def _find_month_columns(ws) -> dict[int, Tuple[int, int]]:
    """Findet die Spaltenpaare für jeden Monat.

    Sucht in Zeile 1 nach Monatsnamen. Der Monatsname steht in der
    Inhalt-Spalte; die Tagesnummern-Spalte ist eine davor.

    Returns:
        Dict: {Monatsnummer: (day_col, event_col)}
    """
    month_cols: dict[int, Tuple[int, int]] = {}

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if isinstance(val, str):
            val_clean = val.strip()
            for i, name in enumerate(MONTH_NAMES):
                if val_clean.lower() == name.lower():
                    month_num = i + 1
                    # Der Monatsname steht in der Inhalt-Spalte
                    # Die Tagesnummern-Spalte ist eine davor
                    month_cols[month_num] = (col - 1, col)
                    break

    return month_cols


def _find_day_rows_for_column(ws, day_col: int) -> dict[int, int]:
    """Findet die Zeilen für Tagesnummern (1-31) in einer bestimmten Spalte.

    Jeder Monat hat seine eigene Tagesnummern-Spalte, da die Monate
    auf verschiedenen Zeilen beginnen und enden.

    Args:
        ws: Worksheet
        day_col: Spaltennummer der Tagesnummern (z.B. 2=B für Jan, 4=D für Feb)

    Returns:
        Dict: {Zeilennummer: Tagesnummer}
    """
    day_rows: dict[int, int] = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=day_col).value
        if val is not None:
            try:
                day = int(val)
                if 1 <= day <= 31:
                    day_rows[row] = day
            except (ValueError, TypeError):
                pass
    return day_rows


def _get_cell_text(cell: Cell) -> str:
    """Extrahiert den Text einer Zelle."""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _get_cell_color(cell: Cell) -> Optional[str]:
    """Extrahiert die Hintergrundfarbe einer Zelle (als Hex)."""
    try:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
            if rgb != "00000000":
                return rgb
    except Exception:
        pass
    return None
